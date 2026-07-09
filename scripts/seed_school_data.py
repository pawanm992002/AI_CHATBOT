"""
Seed school ERP data from an Excel workbook into MongoDB.

Usage:
    python scripts/seed_school_data.py --source-file sample_data/school_erp_sample.xlsx --dev

    --source-file  Path to the Excel file (default: sample_data/school_erp_sample.xlsx)
    --dev          Print credentials to stdout (omitted in non-dev environments)
"""

import argparse
import hashlib
import os
import sys
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

# Ensure backend is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from bson.decimal128 import Decimal128
from core.auth import db, get_password_hash, hash_api_key
from core.config import settings


SCHOOLS_DATA = [
    {
        "school_id": 1,
        "business_name": "ABC Public School",
        "domain": "abc.school.com",
        "email": "abc@school.com",
    },
    {
        "school_id": 2,
        "business_name": "XYZ Convent School",
        "domain": "xyz.school.com",
        "email": "xyz@school.com",
    },
    {
        "school_id": 3,
        "business_name": "Sunrise International Academy",
        "domain": "sunrise.school.com",
        "email": "sunrise@school.com",
    },
]

SHEETS_MAP = {
    "schools": "schools",
    "classes": "school_classes",
    "sections": "school_sections",
    "students": "school_students",
    "routes": "school_routes",
    "stops": "school_stops",
    "transport_assign": "school_transport_assign",
    "hostel_assign": "school_hostel_assign",
    "applied_fees": "school_applied_fees",
    "payments": "school_payments",
}

MONEY_FIELDS = {"amount", "paid_amount", "balance", "concession"}


def _to_decimal(val):
    """Convert int/float to Decimal128 for MongoDB."""
    if val is None:
        return Decimal128("0")
    return Decimal128(str(Decimal(str(val))))


def _transform_row(row: dict, sheet_name: str) -> dict:
    """Convert money fields to Decimal128 and ensure required casting."""
    transformed = {}
    for k, v in row.items():
        if k in MONEY_FIELDS and v is not None:
            transformed[k] = _to_decimal(v)
        else:
            transformed[k] = v
    return transformed


async def seed(dev_mode: bool, source_file: str):
    source_path = Path(source_file)
    if not source_path.exists():
        print(f"[SEED] ERROR: Source file not found: {source_path}")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(source_path, data_only=True)

    tenant_map = {}  # school_id -> tenant_id

    print("[SEED] Creating tenants...")
    for school in SCHOOLS_DATA:
        tenant_id = str(uuid.uuid4())
        api_key = f"sk_live_{uuid.uuid4().hex}"
        password = uuid.uuid4().hex[:12]
        now = datetime.now(timezone.utc)

        existing = await db.tenants.find_one({"email": school["email"]})
        if existing:
            print(f"[SEED] Tenant {school['business_name']} already exists (tenant_id={existing['tenant_id']}), skipping creation.")
            tenant_map[school["school_id"]] = existing["tenant_id"]
            continue

        tenant_doc = {
            "tenant_id": tenant_id,
            "api_key": api_key,
            "api_key_hash": hash_api_key(api_key),
            "domain": school["domain"],
            "business_name": school["business_name"],
            "email": school["email"],
            "password_hash": get_password_hash(password),
            "created_at": now,
            "status": "approved",
            "plan": "school",
            "show_sources": True,
        }
        await db.tenants.insert_one(tenant_doc)
        tenant_map[school["school_id"]] = tenant_id

        creds_file = Path(f".school_credentials_{school['business_name'].replace(' ', '_').lower()}.txt")
        creds_content = (
            f"School: {school['business_name']}\n"
            f"Email:  {school['email']}\n"
            f"Password: {password}\n"
            f"Tenant ID: {tenant_id}\n"
            f"API Key: {api_key}\n"
            f"{'='*40}\n"
        )

        if dev_mode:
            print(f"\n[SEED] === Tenant created: {school['business_name']} ===")
            print(creds_content)
        else:
            creds_file.write_text(creds_content)
            print(f"[SEED] Tenant {school['business_name']} created. Credentials written to {creds_file}")
            print(f"[SEED]   WARNING: Keep this file secure. Delete after recording the credentials.")

    print(f"[SEED] Tenant map: {tenant_map}")

    print("[SEED] Seeding school ERP data...")
    seen_collections = set()

    for sheet_name, coll_name in SHEETS_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"[SEED] Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not headers:
            continue

        coll = db[coll_name]
        seen_collections.add(coll_name)

        total_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict = dict(zip(headers, row))
            school_id = row_dict.get("school_id")
            tenant_id = tenant_map.get(school_id)
            if not tenant_id:
                print(f"[SEED] WARNING: No tenant found for school_id={school_id} in {sheet_name}, skipping row.")
                continue

            row_dict["tenant_id"] = tenant_id
            transformed = _transform_row(row_dict, sheet_name)
            await coll.insert_one(transformed)
            total_rows += 1

        print(f"[SEED] {sheet_name} -> {coll_name}: {total_rows} rows seeded")

    print("[SEED] Done.")

    if not dev_mode and not os.environ.get("APP_ENV") == "development":
        print("[SEED] WARNING: Running outside --dev mode. Ensure credentials are stored securely.")


async def main():
    parser = argparse.ArgumentParser(description="Seed school ERP data into MongoDB")
    parser.add_argument(
        "--source-file",
        default="sample_data/school_erp_sample.xlsx",
        help="Path to the Excel workbook (default: sample_data/school_erp_sample.xlsx)",
    )
    parser.add_argument(
        "--dev",
        action="store_true",
        help="Print credentials to stdout (for development environments only)",
    )
    args = parser.parse_args()
    await seed(dev_mode=args.dev, source_file=args.source_file)


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
